from flask import Blueprint, request, jsonify
import sqlite3
from openpyxl import load_workbook
from io import BytesIO
from log import log_user_action

import_excel_bp = Blueprint('import_excel', __name__)


@import_excel_bp.route('/import_excel/<int:user_id>', methods=['POST'])
def import_excel(user_id):
    # Проверка наличия файла
    if 'file' not in request.files:
        return jsonify({"error": "Файл не был передан"}), 400

    file = request.files['file']

    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "Поддерживаются только .xlsx файлы"}), 400

    try:
        wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if not rows:
            return jsonify({"error": "Файл пустой или не содержит данных"}), 400

        valid_data_found = False
        count = 0
        with sqlite3.connect('storage.db') as conn:
            cursor = conn.cursor()
            for row in rows:
                if not any(row):
                    continue

                name = str(row[0]).strip() if row[0] else ""
                quantity = row[1]
                note = str(row[2]).strip() if row[2] else ""

                if len(name) > 10:
                    continue

                if not isinstance(quantity, (int, float)) or quantity < 0 or quantity > 99999:
                    continue

                valid_data_found = True

                cursor.execute(
                    'INSERT INTO items (user_id, name, quantity, note) VALUES (?, ?, ?, ?)',
                    (user_id, name, quantity, note)
                )
                count += 1
            conn.commit()

        if not valid_data_found:
            return jsonify(
                {"error": "Файл имеет неправильную структуру. Данные не соответствуют требуемым условиям."}), 400

        log_user_action(user_id, f"Выполнен импорт данных файла {file}. Импортировано {count} записей.")

        return jsonify({"message": f"Импортировано {count} записей. Требуется перезапуск."}), 200

    except Exception as e:
        print("Ошибка при импорте:", e)
        return jsonify({"error": "Не удалось обработать файл. Убедитесь, что структура Excel файла корректна."}), 500
